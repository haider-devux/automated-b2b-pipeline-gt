"""
Export the retrieved pipeline data to Excel — the single source of truth for "who we found".

Everything is written as **.xlsx** (opens in Excel 2010 and newer). There are two kinds of file, both
living in exports/:

  1. A per-run SNAPSHOT   ->  exports/granjur_report_<timestamp>.xlsx
     A frozen picture of the whole database at the moment one pipeline run finished. Two sheets:
       * Summary   — how many leads sit at each pipeline stage,
       * Companies — one row per lead with its full detail.
     We keep the most recent few and quietly prune the rest so the folder stays tidy.

  2. ONE central database ->  exports/granjur_central.xlsx
     The master file that accumulates every run's info together. Three sheets:
       * Summary      — current status counts,
       * Runs Log     — one appended row per run/download (timestamp + the stage breakdown that day),
       * Latest Leads — the full current lead list, refreshed live each time.
     This is the file to open when you want "the whole database in one place".

One row per LEAD (joined to its company), mapped to the outreach-ready schema (core lead/contact info,
OpenStreetMap physical/local data, a tech signal, plus pipeline/CRM columns). Columns the FREE sources
don't capture are left blank — no fabricated firmographics.

  python scripts/export_leads_csv.py                 # snapshot + update central (exports/*.xlsx)
  python scripts/export_leads_csv.py --status CONTACTED   # snapshot of one stage only
  python scripts/export_leads_csv.py --out C:/path/my.xlsx
  python scripts/export_leads_csv.py --csv           # also drop a raw .csv (opt-in; off by default)
"""
import argparse
import csv
import importlib.util
import os
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import psycopg2
from psycopg2.extras import RealDictCursor

ROOT = Path(__file__).resolve().parent.parent  # scripts/ -> project root
EXPORT_DIR = ROOT / "exports"
CENTRAL_NAME = "granjur_central.xlsx"          # the one accumulating master file
DISCOVERED_NAME = "discovered.xlsx"            # the raw discovery lake (one row per company)
SNAPSHOT_GLOB = "granjur_report_*.xlsx"        # per-run frozen snapshots
KEEP_SNAPSHOTS = 8                             # prune older snapshots beyond this many


def _load_db_config():
    """Reuse wf3_python/config.py's DB dict — the single place the password/host live."""
    cfg_path = ROOT / "wf3_python" / "config.py"
    spec = importlib.util.spec_from_file_location("wf3_config_for_export", cfg_path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.DB


def get_connection():
    db = _load_db_config()
    return psycopg2.connect(host=db["host"], port=db["port"], dbname=db["dbname"],
                            user=db["user"], password=db["password"])


# The requested outreach schema, in order. These become the sheet / CSV header.
COLUMNS = [
    # Core lead & contact info
    "company_name", "website",
    "contact_email", "contact_title",
    # OpenStreetMap physical & local data
    "osm_address", "city", "country", "osm_category",
    # tech signal
    "tech_stack_mentioned",
    # pipeline / CRM columns
    "region", "pipeline_status", "icp_segment", "email_status", "phone",
    # status journey — first time each stage was reached (derived from lead_events)
    "discovered_at", "enriched_at", "qualified_at", "queued_at", "contacted_at", "replied_at",
    "last_activity",
]

# The raw "discovery lake" schema for discovered.xlsx — one row per COMPANY, every field we captured,
# no segmentation/curation. (companies table, latest lead's status/email joined on.)
DISCOVERED_COLUMNS = [
    "company_id", "company_name", "domain", "website", "phone",
    "region", "country", "city", "coordinates",
    "niche", "source", "discovery_cell",
    "gmaps_rating", "gmaps_reviews",
    "employee_count", "tech_stack", "hiring_signals", "intent_strings",
    "lighthouse_mobile", "funding_stage",
    "first_seen_at", "last_verified_at",
    "status", "email", "email_status",
]

# lead_events.to_status -> the journey column it stamps (first time that stage was reached).
_STAGE_TS_MAP = {
    "DISCOVERED": "discovered_at", "ENRICHED": "enriched_at", "QUALIFIED": "qualified_at",
    "QUEUED_FOR_OUTREACH": "queued_at", "CONTACTED": "contacted_at", "REPLIED": "replied_at",
}

# Leads we actually gathered (skip the internal ERROR/parked rows unless --all is passed).
_REAL_STATUSES = ("DISCOVERED", "ENRICHING", "ENRICHED", "QUALIFYING", "QUALIFIED",
                  "TRANSLATING", "QUEUED_FOR_OUTREACH", "CONTACTED", "REPLIED", "BOOKED",
                  "WON", "NEEDS_CONTACT")

_SQL = """
    SELECT
        c.legal_name, c.domain, c.website_url, c.phone, c.region, c.country, c.city,
        c.niche, c.employee_count, c.tech_stack, c.active_job_posts, c.intent_strings,
        c.raw_payload,
        l.id AS lead_id,
        l.email, l.first_name, l.last_name, l.linkedin_url, l.job_title, l.seniority,
        l.status, l.icp_segment, l.email_validation_status,
        l.pitch_subject, l.pitch_body
    FROM leads l
    JOIN companies c ON c.id = l.company_id
    {where}
    ORDER BY c.region, c.legal_name;
"""

# discovered.xlsx source: one row per COMPANY (all of them), with the latest lead's status/email.
_SQL_COMPANIES = """
    SELECT
        c.id, c.legal_name, c.domain, c.website_url, c.phone, c.region, c.country, c.city,
        c.niche, c.employee_count, c.gmaps_rating, c.gmaps_reviews, c.tech_stack,
        c.lighthouse_mobile, c.active_job_posts, c.intent_strings, c.funding_stage,
        c.source, c.discovery_cell, c.raw_payload, c.first_seen_at, c.last_verified_at,
        l.status, l.email, l.email_validation_status
    FROM companies c
    LEFT JOIN LATERAL (
        SELECT status, email, email_validation_status
        FROM leads WHERE company_id = c.id
        ORDER BY created_at DESC LIMIT 1
    ) l ON true
    ORDER BY c.first_seen_at DESC, c.legal_name;
"""


def _stamp():
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _fmt_ts(ts):
    """Format a timestamp for a cell as a plain string. openpyxl can't store tz-aware datetimes, so we
    always render to text — no timezone crash, and it sorts fine in Excel."""
    if not ts:
        return ""
    try:
        return ts.strftime("%Y-%m-%d %H:%M")
    except AttributeError:
        return str(ts)


def _first_sentence(text):
    if not text:
        return ""
    # skip a leading "Hi X," greeting line, then take the first real sentence.
    body = re.sub(r"^\s*(hi|hello|dear)\b[^\n,]*,\s*", "", text.strip(), flags=re.I)
    body = body.strip().split("\n", 1)[0].strip() or text.strip().split("\n", 1)[0]
    m = re.match(r"(.{20,180}?[.!?])(\s|$)", body)
    return (m.group(1) if m else body[:180]).strip()


def _jobs_list(raw):
    return raw if isinstance(raw, list) else []


def _derive_first_line(row, job):
    """A grounded opening line for the mail-merge. Prefer the real AI pitch's first sentence;
    else build one from the strongest signal we DO have (hiring role, then slow/legacy tech)."""
    fs = _first_sentence(row.get("pitch_body"))
    if fs:
        return fs
    city = (row.get("city") or "").strip()
    if job and job.get("title"):
        where = f" in {city}" if city else ""
        return f"I noticed {row['legal_name']} is hiring a {job['title']}{where} — usually a sign the team is scaling fast."
    tech = row.get("tech_stack") or []
    if tech:
        return f"I had a look at {row['legal_name']}'s site and noticed it's built on {tech[0]} — happy to share one quick win we often find there."
    return ""


def _hiring_pain_point(job, tech):
    if not job or not job.get("title"):
        return ""
    title = job["title"]
    stack = f" (stack mentions {', '.join(tech[:3])})" if tech else ""
    return f"Actively hiring a {title}{stack} — likely under pressure to ship/scale and short on delivery capacity."


def _coordinates(raw):
    if not isinstance(raw, dict):
        return ""
    lat, lon = raw.get("lat"), raw.get("lon")
    if lat is None or lon is None:
        # collect_osm may nest under raw['osm'] in future — check gracefully
        osm = raw.get("osm") if isinstance(raw.get("osm"), dict) else {}
        lat, lon = osm.get("lat"), osm.get("lon")
    return f"{lat},{lon}" if lat is not None and lon is not None else ""


def _osm_address(row, raw):
    if isinstance(raw, dict) and raw.get("address"):
        return raw["address"]
    parts = [row.get("city"), row.get("country")]
    return ", ".join(p for p in parts if p)


def build_row(row, journey=None):
    """One central-file row (lead + company). `journey` is this lead's {stage: datetime} map from
    stage_timestamps(); its columns are rendered as text so Excel never chokes on tz-aware datetimes."""
    journey = journey or {}
    raw = row.get("raw_payload") or {}
    tech = list(row.get("tech_stack") or [])
    jobs = _jobs_list(row.get("active_job_posts"))
    job = jobs[0] if jobs else None
    website = row.get("website_url") or (("https://" + row["domain"]) if row.get("domain") else "")
    base = {
        "company_name": row.get("legal_name") or "",
        "website": website,
        "contact_first_name": row.get("first_name") or "",
        "contact_last_name": row.get("last_name") or "",
        "contact_email": row.get("email") or "",
        "contact_title": row.get("job_title") or "",
        "osm_address": _osm_address(row, raw),
        "city": row.get("city") or "",
        "country": row.get("country") or row.get("region") or "",
        "osm_category": row.get("niche") or "",
        "osm_coordinates": _coordinates(raw),
        "job_title_hiring": (job or {}).get("title", "") if job else (row.get("job_title") or ""),
        "job_board_source": (job or {}).get("source", "") if job else "",
        "job_posting_url": (job or {}).get("url", "") if job else "",
        "tech_stack_mentioned": ";".join(tech),
        "hiring_pain_point": _hiring_pain_point(job, tech),
        "first_line": _derive_first_line(row, job),
        "competitor_name": "",   # not captured by free sources — left blank rather than fabricated
        "region": row.get("region") or "",
        "pipeline_status": row.get("status") or "",
        "icp_segment": row.get("icp_segment") or "",
        "email_status": row.get("email_validation_status") or "",
        "phone": row.get("phone") or "",
    }
    # status journey (first time each stage was reached) — always text via _fmt_ts
    for stage_col in ("discovered_at", "enriched_at", "qualified_at", "queued_at",
                      "contacted_at", "replied_at", "last_activity"):
        base[stage_col] = _fmt_ts(journey.get(stage_col))
    return base


def build_company_row(row):
    """One discovered.xlsx row — the raw discovery record for a company (everything we found)."""
    raw = row.get("raw_payload") or {}
    jobs = _jobs_list(row.get("active_job_posts"))
    website = row.get("website_url") or (("https://" + row["domain"]) if row.get("domain") else "")

    def _n(v):
        return v if v is not None else ""

    return {
        "company_id": str(row.get("id") or ""),
        "company_name": row.get("legal_name") or "",
        "domain": row.get("domain") or "",
        "website": website,
        "phone": row.get("phone") or "",
        "region": row.get("region") or "",
        "country": row.get("country") or "",
        "city": row.get("city") or "",
        "coordinates": _coordinates(raw),
        "niche": row.get("niche") or "",
        "source": row.get("source") or "",
        "discovery_cell": row.get("discovery_cell") or "",
        "gmaps_rating": float(row["gmaps_rating"]) if row.get("gmaps_rating") is not None else "",
        "gmaps_reviews": _n(row.get("gmaps_reviews")),
        "employee_count": _n(row.get("employee_count")),
        "tech_stack": ";".join(list(row.get("tech_stack") or [])),
        "hiring_signals": "; ".join(j.get("title", "") for j in jobs if j.get("title")),
        "intent_strings": ";".join(list(row.get("intent_strings") or [])),
        "lighthouse_mobile": _n(row.get("lighthouse_mobile")),
        "funding_stage": row.get("funding_stage") or "",
        "first_seen_at": _fmt_ts(row.get("first_seen_at")),
        "last_verified_at": _fmt_ts(row.get("last_verified_at")),
        "status": row.get("status") or "",
        "email": row.get("email") or "",
        "email_status": row.get("email_validation_status") or "",
    }


# Pipeline stages in order, with a plain-English label for the Summary sheet.
STATUS_ORDER = [
    ("DISCOVERED", "Discovered — found, not yet processed"),
    ("ENRICHING", "Enriching — being looked up right now"),
    ("ENRICHED", "Enriched — email + tech found"),
    ("QUALIFYING", "Qualifying — being scored"),
    ("QUALIFIED", "Qualified — pitch written, ready to approve"),
    ("QUEUED_FOR_OUTREACH", "Queued — approved, ready to send"),
    ("CONTACTED", "Contacted — first email sent"),
    ("REPLIED", "Replied — prospect answered"),
    ("BOOKED", "Booked — call scheduled"),
    ("NEEDS_CONTACT", "Needs contact — qualified but no personal email"),
    ("COOLDOWN", "Cooldown — skipped, resting before re-review"),
    ("SUPPRESSED", "Suppressed — bounced/unsubscribed, never mailed again"),
    ("ERROR", "Error — no usable email or data"),
]

# The Runs Log columns (Runs Log sheet in the central file) — one appended row per run/download.
RUN_LOG_HEADERS = ["Run (UTC)", "Total leads", "Discovered", "Enriched", "Qualified",
                   "Queued (send-ready)", "Contacted", "Needs contact", "Replied",
                   "Booked", "Error", "Snapshot file"]


def status_counts(conn):
    with conn.cursor() as cur:
        cur.execute("SELECT status::text, count(*) FROM leads GROUP BY status;")
        return dict(cur.fetchall())


def stage_timestamps(conn):
    """Return {lead_id(str): {enriched_at: dt, queued_at: dt, ..., last_activity: dt}} — the FIRST time
    each lead reached each tracked stage (plus its most recent activity), from the lead_events log."""
    out = {}
    with conn.cursor() as cur:
        cur.execute("""SELECT lead_id, to_status::text, MIN(at)
                         FROM lead_events
                        WHERE to_status::text = ANY(%s)
                        GROUP BY lead_id, to_status;""", (list(_STAGE_TS_MAP.keys()),))
        for lead_id, status, ts in cur.fetchall():
            col = _STAGE_TS_MAP.get(status)
            if col and lead_id is not None:
                out.setdefault(str(lead_id), {})[col] = ts
        cur.execute("SELECT lead_id, MAX(at) FROM lead_events WHERE lead_id IS NOT NULL GROUP BY lead_id;")
        for lead_id, ts in cur.fetchall():
            out.setdefault(str(lead_id), {})["last_activity"] = ts
    return out


def fetch_rows_and_counts(conn=None):
    """Pull every lead (all statuses) + per-status counts + the journey timestamps. Returns (rows, counts)."""
    close = False
    if conn is None:
        conn = get_connection()
        close = True
    try:
        counts = status_counts(conn)
        journey = stage_timestamps(conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(_SQL.format(where=""))          # ALL leads, every status
            rows = [build_row(r, journey.get(str(r.get("lead_id")))) for r in cur.fetchall()]
    finally:
        if close:
            conn.close()
    return rows, counts


def fetch_company_rows(conn=None):
    """Pull EVERY company (raw discovery lake) for discovered.xlsx. Returns a list of built rows."""
    close = False
    if conn is None:
        conn = get_connection()
        close = True
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(_SQL_COMPANIES)
            return [build_company_row(r) for r in cur.fetchall()]
    finally:
        if close:
            conn.close()


# ----------------------------------------------------------------- styling helpers
def _styles():
    from openpyxl.styles import Font, PatternFill
    return {
        "head_fill": PatternFill("solid", fgColor="1F3864"),
        "head_font": Font(bold=True, color="FFFFFF"),
        "title_font": Font(bold=True, size=14),
        "bold": Font(bold=True),
    }


def _fill_summary_sheet(ws, counts, subtitle=""):
    st = _styles()
    ws["A1"] = "Granjur B2B Pipeline — Lead Report"; ws["A1"].font = st["title_font"]
    ws["A2"] = (f"Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}   ·   "
                f"{sum(counts.values())} leads total" + (f"   ·   {subtitle}" if subtitle else ""))
    r = 4
    for c, label in ((1, "STATUS"), (2, "COUNT"), (3, "WHAT IT MEANS")):
        cell = ws.cell(r, c, label); cell.font = st["head_font"]; cell.fill = st["head_fill"]
    r += 1
    for status, label in STATUS_ORDER:
        n = counts.get(status, 0)
        if n == 0 and status in ("ENRICHING", "QUALIFYING", "COOLDOWN", "REPLIED", "BOOKED", "SUPPRESSED"):
            continue                                    # hide transient/empty stages to keep it clean
        ws.cell(r, 1, status)
        ws.cell(r, 2, n)
        ws.cell(r, 3, label.split("—", 1)[-1].strip())
        r += 1
    ws.cell(r, 1, "TOTAL").font = st["bold"]
    ws.cell(r, 2, sum(counts.values())).font = st["bold"]
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 52


def _fill_companies_sheet(ws, rows):
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    st = _styles()
    ws.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(1, c); cell.font = st["head_font"]; cell.fill = st["head_fill"]
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append([row.get(k, "") for k in COLUMNS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
    widths = {"company_name": 26, "website": 30, "contact_email": 30, "osm_address": 24,
              "tech_stack_mentioned": 20, "pipeline_status": 20, "osm_category": 18,
              "discovered_at": 16, "enriched_at": 16, "qualified_at": 16, "queued_at": 16,
              "contacted_at": 16, "replied_at": 16, "last_activity": 16}
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 14)


# ----------------------------------------------------------------- discovered.xlsx (raw discovery lake)
def _fill_discovered_sheet(ws, rows):
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    st = _styles()
    ws.append(DISCOVERED_COLUMNS)
    for c in range(1, len(DISCOVERED_COLUMNS) + 1):
        cell = ws.cell(1, c); cell.font = st["head_font"]; cell.fill = st["head_fill"]
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append([row.get(k, "") for k in DISCOVERED_COLUMNS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DISCOVERED_COLUMNS))}{len(rows) + 1}"
    widths = {"company_id": 38, "company_name": 26, "domain": 22, "website": 30, "email": 28,
              "tech_stack": 22, "hiring_signals": 26, "intent_strings": 20, "discovery_cell": 18,
              "first_seen_at": 17, "last_verified_at": 17, "status": 18, "coordinates": 18}
    for i, col in enumerate(DISCOVERED_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 13)


def _fill_by_source_sheet(ws, rows):
    """Small summary: how many companies each bot/source contributed — the at-a-glance fleet scoreboard."""
    st = _styles()
    ws["A1"] = "Discovered — by source (which bot found it)"; ws["A1"].font = st["title_font"]
    ws["A2"] = f"Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}   ·   {len(rows)} compan(ies) total"
    counts = Counter((r.get("source") or "unknown") for r in rows)
    r = 4
    for c, label in ((1, "SOURCE"), (2, "COMPANIES")):
        cell = ws.cell(r, c, label); cell.font = st["head_font"]; cell.fill = st["head_fill"]
    r += 1
    for src, n in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws.cell(r, 1, src)
        ws.cell(r, 2, n)
        r += 1
    ws.cell(r, 1, "TOTAL").font = st["bold"]
    ws.cell(r, 2, sum(counts.values())).font = st["bold"]
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14


def write_discovered(rows, out_path=None):
    """Write the raw discovery lake (discovered.xlsx): a Discovered sheet (one row per company, every
    field) + a By Source summary. Live DB mirror — rebuilt each run. Returns (path, n_rows).
    Swallows a save error (e.g. the file open in Excel) so it never crashes the pipeline run."""
    from openpyxl import Workbook
    out_path = Path(out_path or (EXPORT_DIR / DISCOVERED_NAME))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _fill_discovered_sheet(wb.active, rows)
    wb.active.title = "Discovered"
    _fill_by_source_sheet(wb.create_sheet("By Source"), rows)
    try:
        wb.save(out_path)
    except Exception:  # noqa: BLE001 — locked open in Excel etc.; don't fail the run
        pass
    return out_path, len(rows)


# ----------------------------------------------------------------- snapshot (per-run frozen file)
def write_snapshot(rows, counts, out_path):
    """Write a standalone snapshot workbook (Summary + Companies) to out_path."""
    from openpyxl import Workbook
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _fill_summary_sheet(wb.active, counts)
    wb.active.title = "Summary"
    _fill_companies_sheet(wb.create_sheet("Companies"), rows)
    wb.save(out_path)
    return out_path


def _prune_snapshots(keep=KEEP_SNAPSHOTS):
    """Keep only the most recent `keep` per-run snapshots so exports/ stays tidy."""
    snaps = sorted(EXPORT_DIR.glob(SNAPSHOT_GLOB))
    for old in snaps[:-keep] if keep > 0 else snaps:
        try:
            old.unlink()
        except OSError:
            pass


# ----------------------------------------------------------------- central (accumulating master)
def _run_log_row(counts, stamp, snapshot_name):
    g = lambda k: counts.get(k, 0)   # noqa: E731
    return [stamp, sum(counts.values()), g("DISCOVERED"), g("ENRICHED"), g("QUALIFIED"),
            g("QUEUED_FOR_OUTREACH"), g("CONTACTED"), g("NEEDS_CONTACT"), g("REPLIED"),
            g("BOOKED"), g("ERROR"), snapshot_name or ""]


def _new_runs_log_sheet(wb):
    """Create a freshly-styled (empty) Runs Log sheet with the header row. Returns the sheet."""
    from openpyxl.styles import Font
    rl = wb.create_sheet("Runs Log")
    rl.append(RUN_LOG_HEADERS)
    for c in range(1, len(RUN_LOG_HEADERS) + 1):
        rl.cell(1, c).font = Font(bold=True)
    rl.freeze_panes = "A2"
    for i, w in enumerate([20, 11, 11, 10, 10, 18, 11, 13, 9, 8, 8, 32], start=1):
        rl.column_dimensions[chr(64 + i)].width = w
    return rl


def _stamp_from_snapshot_name(name):
    """Pull the 20260717_065550 timestamp out of a snapshot filename (fallback: the name itself)."""
    m = re.search(r"(\d{8}_\d{6})", name)
    return m.group(1) if m else name


def _counts_from_snapshot(path):
    """Read a per-run snapshot's Summary sheet back into a {STATUS: count} dict, so a Runs Log row can
    be reconstructed from the file's OWN numbers (they can never drift from what the snapshot shows)."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Summary"]
        counts = {}
        for row in ws.iter_rows(min_row=5, values_only=True):   # status table starts at row 5
            if not row or row[0] in (None, "", "TOTAL"):
                continue
            try:
                counts[str(row[0]).strip()] = int(row[1]) if row[1] is not None else 0
            except (ValueError, TypeError):
                continue
        return counts
    finally:
        wb.close()


def _reorder_sheets(wb, order):
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else len(order))


def refresh_central(rows, counts, stamp, snapshot_name=None, add_run_row=False,
                    central_path=None, return_bytes=False):
    """Update the one central master file: refresh Summary + Latest Leads from the live DB and
    (optionally) append one row to the Runs Log. Saves to disk when it can; if the file is open in
    Excel (locked) it keeps going so a download still succeeds. Returns the path, or bytes when asked."""
    from openpyxl import Workbook, load_workbook
    central_path = Path(central_path or (EXPORT_DIR / CENTRAL_NAME))
    central_path.parent.mkdir(parents=True, exist_ok=True)

    # Load the existing central file (to preserve the Runs Log history) or start fresh.
    if central_path.exists():
        try:
            wb = load_workbook(central_path)
        except Exception:  # noqa: BLE001 — corrupt/locked-open -> rebuild from scratch
            wb = Workbook()
            wb.remove(wb.active)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # ---- Runs Log: keep prior rows, append this run if asked ----
    if "Runs Log" not in wb.sheetnames:
        _new_runs_log_sheet(wb)
    if add_run_row:
        wb["Runs Log"].append(_run_log_row(counts, stamp, snapshot_name))

    # ---- Summary + Latest Leads: always rebuilt from the live DB ----
    for name in ("Summary", "Latest Leads"):
        if name in wb.sheetnames:
            del wb[name]
    _fill_summary_sheet(wb.create_sheet("Summary"), counts, subtitle="central database")
    _fill_companies_sheet(wb.create_sheet("Latest Leads"), rows)

    _reorder_sheets(wb, ["Summary", "Runs Log", "Latest Leads"])

    try:
        wb.save(central_path)          # may fail if the user has it open in Excel — that's fine
    except Exception:                  # noqa: BLE001
        pass
    if return_bytes:
        bio = BytesIO(); wb.save(bio); return bio.getvalue()
    return central_path


def rebuild_runs_log(conn=None, central_path=None):
    """Reconstruct the Runs Log so EVERY row matches an actual snapshot file's own numbers, in time
    order. Reads each granjur_report_<ts>.xlsx's Summary sheet, so the log can never drift from the
    files it references; it also back-fills rows for snapshots that predate the central file, and drops
    any stale/duplicate rows. Summary + Latest Leads are refreshed from the live DB. Returns row count."""
    from openpyxl import Workbook, load_workbook
    central_path = Path(central_path or (EXPORT_DIR / CENTRAL_NAME))
    central_path.parent.mkdir(parents=True, exist_ok=True)

    # one Runs Log row per snapshot file present, using THAT file's own Summary numbers
    runs = []
    for snap in sorted(EXPORT_DIR.glob(SNAPSHOT_GLOB)):
        try:
            counts = _counts_from_snapshot(snap)
        except Exception:  # noqa: BLE001 — skip an unreadable snapshot rather than fail the rebuild
            continue
        runs.append(_run_log_row(counts, _stamp_from_snapshot_name(snap.name), snap.name))

    rows, counts_now = fetch_rows_and_counts(conn)

    if central_path.exists():
        try:
            wb = load_workbook(central_path)
        except Exception:  # noqa: BLE001
            wb = Workbook(); wb.remove(wb.active)
    else:
        wb = Workbook(); wb.remove(wb.active)

    for name in ("Summary", "Runs Log", "Latest Leads"):
        if name in wb.sheetnames:
            del wb[name]
    rl = _new_runs_log_sheet(wb)
    for r in runs:
        rl.append(r)
    _fill_summary_sheet(wb.create_sheet("Summary"), counts_now, subtitle="central database")
    _fill_companies_sheet(wb.create_sheet("Latest Leads"), rows)
    _reorder_sheets(wb, ["Summary", "Runs Log", "Latest Leads"])
    wb.save(central_path)
    return len(runs)


def central_bytes(conn=None):
    """Live bytes of the central database for the dashboard download (does NOT append a Runs Log row —
    a plain download isn't a pipeline run). Refreshes Summary + Latest Leads from the current DB."""
    rows, counts = fetch_rows_and_counts(conn)
    return refresh_central(rows, counts, _stamp(), add_run_row=False, return_bytes=True)


# ----------------------------------------------------------------- top-level export
def export_xlsx(out_path=None, conn=None, central=True):
    """The pipeline-run export. Writes a dated per-run snapshot AND folds it into the central master
    (append a Runs Log row + refresh Summary/Latest Leads), then prunes old snapshots.
    Returns (snapshot_path, n_rows). Signature kept stable for run_pipeline.py + the dashboard."""
    rows, counts = fetch_rows_and_counts(conn)
    stamp = _stamp()
    if out_path is None:
        out_path = EXPORT_DIR / f"granjur_report_{stamp}.xlsx"
    snap = write_snapshot(rows, counts, out_path)

    # discovered.xlsx — the raw discovery lake (one row per company, live DB mirror). Written EVERY run,
    # independent of the central flag. Never let it crash the run (its own save is already guarded).
    try:
        drows = fetch_company_rows(conn)
        dpath, dn = write_discovered(drows)
        print(f"  Wrote discovery lake: {dn} compan(ies) -> {dpath}")
    except Exception as e:  # noqa: BLE001
        print(f"  discovered.xlsx export failed: {e}")

    if central:
        refresh_central(rows, counts, stamp, snapshot_name=snap.name, add_run_row=True)
        _prune_snapshots()
    return snap, len(rows)


# ----------------------------------------------------------------- optional raw CSV (opt-in)
def export_csv(out_path=None, status=None, include_all=False, conn=None):
    """Opt-in raw CSV dump (off by default — we standardised on .xlsx). Same columns as the sheets."""
    close = False
    if conn is None:
        conn = get_connection()
        close = True
    try:
        if status:
            where, params = "WHERE l.status = %s::lead_status", (status,)
        elif include_all:
            where, params = "", ()
        else:
            where, params = "WHERE l.status::text = ANY(%s)", (list(_REAL_STATUSES),)
        journey = stage_timestamps(conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(_SQL.format(where=where), params)
            rows = cur.fetchall()
    finally:
        if close:
            conn.close()

    if out_path is None:
        out_path = EXPORT_DIR / f"granjur_leads_{_stamp()}.csv"
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")  # build_row has extra keys; drop them
        w.writeheader()
        for r in rows:
            w.writerow(build_row(r, journey.get(str(r.get("lead_id")))))
    return out_path, len(rows)


def main():
    ap = argparse.ArgumentParser(description="Export retrieved pipeline leads to Excel (snapshot + central).")
    ap.add_argument("--out", metavar="PATH", help="explicit snapshot .xlsx path (default exports/granjur_report_<ts>.xlsx)")
    ap.add_argument("--no-central", action="store_true", help="write only the snapshot; don't touch the central file")
    ap.add_argument("--rebuild-log", action="store_true", help="rebuild the central Runs Log from the snapshot "
                    "files present (each row matches that file's own numbers) — repairs drift; writes no new snapshot")
    ap.add_argument("--csv", action="store_true", help="ALSO write a raw .csv (opt-in; .xlsx is the default format)")
    ap.add_argument("--status", metavar="STAGE", help="(--csv only) limit the CSV to one pipeline status")
    ap.add_argument("--all", action="store_true", help="(--csv only) include internal ERROR/parked rows too")
    args = ap.parse_args()

    if args.rebuild_log:
        try:
            n = rebuild_runs_log()
            print(f"Rebuilt central Runs Log from {n} snapshot file(s) -> {EXPORT_DIR / CENTRAL_NAME}")
        except Exception as e:  # noqa: BLE001
            print(f"Runs Log rebuild failed: {e}")
            sys.exit(1)
        return

    try:
        path, n = export_xlsx(out_path=args.out, central=not args.no_central)
        print(f"Wrote snapshot: {n} compan(ies) -> {path}  (opens in Excel 2010)")
        if not args.no_central:
            print(f"Updated central database -> {EXPORT_DIR / CENTRAL_NAME}")
    except Exception as e:  # noqa: BLE001
        print(f"Excel export failed: {e}")
        sys.exit(1)

    if args.csv:
        try:
            cpath, cn = export_csv(status=args.status, include_all=args.all)
            print(f"Also wrote raw CSV: {cn} lead(s) -> {cpath}")
        except Exception as e:  # noqa: BLE001
            print(f"CSV export skipped: {e}")


if __name__ == "__main__":
    main()
